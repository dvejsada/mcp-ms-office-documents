import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import SheetTitleException

from upload_tools import upload_file
from .helpers import add_table_to_sheet
from .parser import (
    walk_markdown_lines,
    collect_table_positions,
    SheetEvent,
    HeaderEvent,
    TableEvent,
    DEFAULT_SHEET_NAME,
    _sanitize_sheet_name,
)

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
# Header font styles by level
HEADER_FONTS = {
    1: Font(size=16, bold=True, color="2F5597"),
    2: Font(size=14, bold=True, color="4472C4"),
}
HEADER_FONT_DEFAULT = Font(size=12, bold=True)


def _warn_on_circular_references(xlsx_bytes: bytes, sheet_names: list[str]) -> None:
    """Log a warning if the saved workbook contains circular references.

    Purely diagnostic — a cycle makes Excel show a warning dialog and resolve
    the cells to 0, which is silent from this server's side, so we surface it
    in the logs. Never raises: a detector failure must not block delivery of
    an otherwise valid document.
    """
    try:
        from .circular_refs import detect_circular_references
        detect_circular_references(xlsx_bytes, sheet_names)
    except Exception as e:  # pragma: no cover — defensive
        logger.debug("Circular-reference detection unavailable: %s", e)


def markdown_to_excel(markdown_content: str, file_name: str | None = None, auto_filter: bool = False) -> str:
    """Convert Markdown to Excel workbook (focused on tables and headers).

    Always starts from an empty Workbook (no templates).
    Supports multiple sheets via '## Sheet: Name' headings.
    Supports cross-sheet references via ``SheetName!T1.B[0]`` syntax.

    Args:
        markdown_content: Markdown string with tables.
        file_name: Optional custom filename (without extension).
        auto_filter: If True, apply Excel auto-filter to each table.

    Raises:
        RuntimeError: If the markdown contains no tables or conversion fails.
    """
    logger.info("Starting markdown_to_excel conversion")

    # ── Input validation ──
    if not markdown_content or not markdown_content.strip():
        raise RuntimeError("Cannot create Excel workbook: markdown content is empty")

    # Split content into lines and parse into events (single shared state machine)
    lines: list[str] = markdown_content.split('\n')
    events = walk_markdown_lines(lines)

    # Build table position map from events (used for cross-sheet formula resolution)
    all_sheet_table_positions = collect_table_positions(events)
    logger.debug("Table positions (all sheets): %s", all_sheet_table_positions)

    # ── Build the actual workbook from events ──
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(DEFAULT_SHEET_NAME)

    # Per-sheet state for formula resolution
    table_positions: dict[str, int] = {}

    # Counters for summary
    headers_count = 0
    tables_count = 0

    # Worksheet titles as openpyxl actually stored them. When a '## Sheet:'
    # name collides with an existing sheet — two long names that become equal
    # after the 31-char truncation, or a plain duplicate — openpyxl silently
    # appends a suffix. The cross-sheet position map is keyed by the requested
    # name, so that divergence routes formulas to the wrong sheet with no
    # error anywhere. Warn so the caller knows to rename.
    seen_sheet_titles: set[str] = {ws.title}

    try:
        for event in events:
            if isinstance(event, SheetEvent):
                if event.is_rename:
                    try:
                        ws.title = event.sheet_name
                        seen_sheet_titles = {ws.title}
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Could not rename worksheet to '%s': %s — using default",
                            event.sheet_name, exc,
                        )
                else:
                    if event.sheet_name in seen_sheet_titles:
                        logger.warning(
                            "Sheet name '%s' collides with an existing sheet after "
                            "sanitization; openpyxl will auto-rename it, which "
                            "breaks cross-sheet references that use the original "
                            "name. Use a distinct sheet name.",
                            event.sheet_name,
                        )
                    try:
                        ws = wb.create_sheet(title=event.sheet_name)
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Invalid sheet name '%s': %s — using fallback",
                            event.sheet_name, exc,
                        )
                        ws = wb.create_sheet()
                    seen_sheet_titles.add(ws.title)
                    table_positions = {}

            elif isinstance(event, HeaderEvent):
                cell = ws.cell(row=event.row, column=1)
                cell.value = event.text
                cell.font = HEADER_FONTS.get(event.level, HEADER_FONT_DEFAULT)
                headers_count += 1
                logger.debug("Header (level %d) at row %d: %s", event.level, event.row, event.text)

            elif isinstance(event, TableEvent):
                # Record this table's position for local formula resolution
                table_positions[event.table_key] = event.start_row

                # Write table to worksheet
                add_table_to_sheet(
                    event.table_data, ws, event.start_row, table_positions,
                    all_sheet_table_positions=all_sheet_table_positions,
                    auto_filter=auto_filter,
                    table_index=tables_count,
                    directives=event.directives,
                )

                # Handle freeze directive — freeze below header row of this table
                if 'freeze' in event.directives:
                    ws.freeze_panes = f"A{event.start_row + 1}"

                tables_count += 1
                logger.debug(
                    "Added table #%d (%s) with %d data rows on sheet '%s'",
                    tables_count, event.table_key, len(event.table_data) - 1, event.sheet_name,
                )

    except Exception as e:
        logger.error("Error generating Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error generating Excel workbook: {e}") from e

    # ── Validation: ensure at least one table was created ──
    if tables_count == 0:
        raise RuntimeError(
            "Cannot create Excel workbook: no valid markdown tables found in the input. "
            "Tables must use pipe syntax (| col1 | col2 |) with a separator row (|---|---|)."
        )

    # Save workbook to BytesIO and upload via existing helper
    file_object = io.BytesIO()
    try:
        logger.info("Saving Excel workbook to memory buffer (headers=%d, tables=%d)", headers_count, tables_count)
        wb.save(file_object)
        _warn_on_circular_references(file_object.getvalue(), wb.sheetnames)
        file_object.seek(0)
        result = upload_file(file_object, "xlsx", filename=file_name)
        logger.info("Excel upload completed successfully")
        return result
    except Exception as e:
        logger.error("Error saving/uploading Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error saving/uploading Excel workbook: {e}") from e
    finally:
        file_object.close()


def _markdown_to_excel_buffer(markdown_content: str, auto_filter: bool = False) -> io.BytesIO:
    """Convert Markdown to Excel workbook and return as BytesIO buffer.

    This function is useful when the caller needs to handle upload separately,
    such as for LibreChat file artifact uploads.

    Args:
        markdown_content: Markdown string with tables.
        auto_filter: If True, apply Excel auto-filter to each table.

    Returns:
        BytesIO buffer containing the Excel workbook (position at start)

    Raises:
        RuntimeError: If the markdown contains no tables or conversion fails.
    """
    logger.info("Starting _markdown_to_excel_buffer conversion")

    # ── Input validation ──
    if not markdown_content or not markdown_content.strip():
        raise RuntimeError("Cannot create Excel workbook: markdown content is empty")

    # Split content into lines and parse into events (single shared state machine)
    lines: list[str] = markdown_content.split('\n')
    events = walk_markdown_lines(lines)

    # Build table position map from events (used for cross-sheet formula resolution)
    all_sheet_table_positions = collect_table_positions(events)

    # ── Build the actual workbook from events ──
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(DEFAULT_SHEET_NAME)

    # Per-sheet state for formula resolution
    table_positions: dict[str, int] = {}

    # Counters for summary
    tables_count = 0

    try:
        for event in events:
            if isinstance(event, SheetEvent):
                if event.is_rename:
                    try:
                        ws.title = event.sheet_name
                    except (SheetTitleException, ValueError):
                        pass
                else:
                    try:
                        ws = wb.create_sheet(title=event.sheet_name)
                    except (SheetTitleException, ValueError):
                        ws = wb.create_sheet()
                    table_positions = {}

            elif isinstance(event, HeaderEvent):
                cell = ws.cell(row=event.row, column=1)
                cell.value = event.text
                cell.font = HEADER_FONTS.get(event.level, HEADER_FONT_DEFAULT)

            elif isinstance(event, TableEvent):
                table_positions[event.table_key] = event.start_row
                add_table_to_sheet(
                    event.table_data, ws, event.start_row, table_positions,
                    all_sheet_table_positions=all_sheet_table_positions,
                    auto_filter=auto_filter,
                    table_index=tables_count,
                    directives=event.directives,
                )
                if 'freeze' in event.directives:
                    ws.freeze_panes = f"A{event.start_row + 1}"
                tables_count += 1

    except Exception as e:
        raise RuntimeError(f"Error generating Excel workbook: {e}") from e

    if tables_count == 0:
        raise RuntimeError(
            "Cannot create Excel workbook: no valid markdown tables found in the input."
        )

    # Save workbook to BytesIO
    file_object = io.BytesIO()
    try:
        wb.save(file_object)
        _warn_on_circular_references(file_object.getvalue(), wb.sheetnames)
        file_object.seek(0)
        return file_object
    except Exception as e:
        raise RuntimeError(f"Error saving Excel workbook: {e}") from e
